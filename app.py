from flask import Flask,render_template,request,url_for,redirect,make_response,flash
from flask_mysqldb import MySQL
from fpdf import FPDF
from io import BytesIO
import openpyxl
from flask_paginate import Pagination, get_page_parameter


app = Flask(__name__)
app.secret_key = "flash message"

app.config['MYSQL_HOST'] ="localhost"
app.config['MYSQL_USER'] ="root"
app.config['MYSQL_PASSWORD'] =""
app.config['MYSQL_DB'] ="azza"

mysql = MySQL(app)


@app.route('/')
def index():
    return render_template('index.html')

######### ADD ##########

@app.route('/add_from_produit')
def add_from_produit():
    return render_template('add_from_produit.html')

@app.route('/add_from_magasin')
def add_from_magasin():
    return render_template('add_from_magasin.html')

@app.route('/add_from_stock')
def add_from_stock():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM produit")
    produit = cur.fetchall()
    cur.execute("SELECT * FROM magasin")
    magasin = cur.fetchall()
    cur.close()
    return render_template('add_from_stock.html', magasin=magasin, produit=produit)

#############################################################################################

#### produit ####

@app.route('/produit')
def list_produit():
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 1
    offset = (page - 1) * per_page
    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) FROM produit")
    total_count = cur.fetchone()[0]
    cur.execute("SELECT * FROM produit LIMIT %s OFFSET %s", (per_page, offset))
    data = cur.fetchall()
    cur.close()
    pagination = Pagination(page=page, total=total_count, per_page=per_page, css_framework='bootstrap5')
    return render_template('list_produit.html', produit=data, pagination=pagination)

@app.route('/add_produit', methods=["POST"])
def add_produit():
    # Obtenir les données du formulaire
    nom = request.form['nom']
    prix = request.form['prix']
    description = request.form['description']

    # Établir une connexion à la base de données
    cur = mysql.connection.cursor()

    # Insérer les données dans la table 'produit'
    cur.execute("INSERT INTO produit (nom, prix, description) VALUES (%s, %s, %s)", (nom, prix, description))

    # Valider la transaction
    mysql.connection.commit()

    # Fermer la connexion à la base de données
    cur.close()

    # Rediriger vers la page 'produit'
    return redirect(url_for('list_produit'))

@app.route('/upd_produit/<int:id>', methods=["GET", "POST"])
def upd_produit(id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM produit WHERE id=%s", (id,))
    produit = cur.fetchone()
    cur.close()
    
    if request.method == "POST":
        nom = request.form['nom']
        prix = request.form['prix']
        description = request.form['description']
        
        cur = mysql.connection.cursor()
        cur.execute("UPDATE produit SET nom=%s, prix=%s, description=%s WHERE id=%s", (nom, prix, description, id))
        mysql.connection.commit()
        cur.close()
        
        return redirect(url_for('list_produit'))
    
    print("Produit : ", produit)
    if produit:
        return render_template('upd_produit.html', produit=produit)
    else:
        return "Produit introuvable"


@app.route('/delete_produit/<string:id_data>', methods=["GET"])
def delete_produit(id_data):
    cur = mysql.connection.cursor()
    cur.execute("SELECT id_produit FROM stock WHERE id_produit=%s", [id_data])
    stock_data = cur.fetchone()
    if stock_data:
        # si l'id existe dans la table stock, ne supprime pas la ligne correspondante dans la table produit
        flash("Impossible de supprimer le produit car il est présent dans le stock")
        return redirect(url_for('list_produit'))
    else:
        cur.execute("DELETE FROM produit WHERE id=%s", [id_data])
        mysql.connection.commit()
        cur.close()
        flash("Le produit a été supprimé avec succès")
        return redirect(url_for('list_produit'))


@app.route('/export_produit_pdf')
def export_produit_pdf():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM produit")
    data = cur.fetchall()
    cur.close()

    # création du rapport en PDF
    pdf = FPDF()
    pdf.add_page()

    # ajout de l'en-tête
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Liste des produit', 0, 1, 'C')
    pdf.set_font('Arial', '', 10)
    pdf.cell(0, 10, 'Ce document contient la liste de tous les produits', 0, 1, 'C')

    # ajout des titres des colonnes
    pdf.set_font('Arial', 'B', 16)
    col_width = pdf.w / 5.5
    row_height = pdf.font_size * 2
    for header in cur.description:
        pdf.cell(col_width, row_height, str(header[0]), border=1)

    # ajout des données de la base de données
    pdf.set_font('Arial', '', 10)
    for row in data:
        pdf.ln()
        for item in row:
            pdf.cell(col_width, row_height, str(item), border=1)

    # conversion du rapport en PDF en octets
    output = BytesIO()
    pdf.output(output)
    pdf_bytes = output.getvalue()

    # envoi du rapport en PDF au navigateur
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=magasins.pdf'
    return response

@app.route('/export_produit_excel')
def export_produit_excel():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM produit")
    data = cur.fetchall()
    cur.close()

    # création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # ajout des titres des colonnes
    headers = [header[0] for header in cur.description]
    for col in range(len(headers)):
        ws.cell(row=1, column=col+1, value=headers[col])

    # ajout des données de la base de données
    for row in range(len(data)):
        for col in range(len(headers)):
            ws.cell(row=row+2, column=col+1, value=data[row][col])

    # conversion du classeur en octets
    output = BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()

    # envoi du classeur Excel au navigateur
    response = make_response(excel_bytes)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=produits.xlsx'
    return response


#### magasin ####

@app.route('/magasin')
def listemagasin():
    page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 5
    offset = (page - 1) * per_page
    cur = mysql.connection.cursor()
    cur.execute("SELECT COUNT(*) FROM magasin")
    total_count = cur.fetchone()[0]
    cur.execute("SELECT * FROM magasin LIMIT %s OFFSET %s", (per_page, offset))
    data = cur.fetchall()
    cur.close()
    pagination = Pagination(page=page, total=total_count, per_page=per_page, css_framework='bootstrap5')
    return render_template('list_magasin.html', magasin=data, pagination=pagination)

@app.route('/add_magasin', methods=["POST"])
def add_magasin():
    # Obtenir les données du formulaire
    nom = request.form['nom']
    adresse = request.form['adresse']

    # Établir une connexion à la base de données
    cur = mysql.connection.cursor()

    # Insérer les données dans la table 'magasin'
    cur.execute("INSERT INTO magasin (nom, adresse) VALUES (%s, %s)", (nom, adresse))

    # Valider la transaction
    mysql.connection.commit()

    # Fermer la connexion à la base de données
    cur.close()

    # Rediriger vers la page 'magasin'
    return redirect(url_for('listemagasin'))

@app.route('/upd_magasin/<int:id>', methods=["GET", "POST"])
def upd_magasin(id):
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM magasin WHERE id=%s", (id,))
    magasin = cur.fetchone()
    cur.close()
    
    if request.method == "POST":
        nom = request.form['nom']
        adresse = request.form['adresse']
        
        cur = mysql.connection.cursor()
        cur.execute("UPDATE magasin SET nom=%s, adresse=%s WHERE id=%s", (nom, adresse, id))
        mysql.connection.commit()
        cur.close()
        
        return redirect(url_for('listemagasin'))

    return render_template('upd_magasin.html', magasin=magasin)

@app.route('/delete_magasin/<string:id_data>', methods=["GET"])
def delete_magasin(id_data):
    cur = mysql.connection.cursor()
    cur.execute("SELECT id_magasin FROM stock WHERE id_magasin=%s", [id_data])
    stock_data = cur.fetchone()
    if stock_data:
        # si l'id existe dans la table stock, ne supprime pas la ligne correspondante dans la table magasin
        flash("Impossible de supprimer le magasin car il est présent dans le stock")
        return redirect(url_for('listemagasin'))
    else:
        cur.execute("DELETE FROM magasin WHERE id=%s", [id_data])
        mysql.connection.commit()
        cur.close()
        flash("Le magasin a été supprimé avec succès")
        return redirect(url_for('listemagasin'))

@app.route('/export_magasin_pdf')
def export_magasin_pdf():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM magasin")
    data = cur.fetchall()
    cur.close()

    # création du rapport en PDF
    pdf = FPDF()
    pdf.add_page()

    # ajout de l'en-tête
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Liste des magasins', 0, 1, 'C')
    pdf.set_font('Arial', '', 10)
    pdf.cell(0, 10, 'Ce document contient la liste de tous les magasins', 0, 1, 'C')

    # ajout des titres des colonnes
    pdf.set_font('Arial', 'B', 16)
    col_width = pdf.w / 5.5
    row_height = pdf.font_size * 2
    for header in cur.description:
        pdf.cell(col_width, row_height, str(header[0]), border=1)

    # ajout des données de la base de données
    pdf.set_font('Arial', '', 10)
    for row in data:
        pdf.ln()
        for item in row:
            pdf.cell(col_width, row_height, str(item), border=1)

    # conversion du rapport en PDF en octets
    output = BytesIO()
    pdf.output(output)
    pdf_bytes = output.getvalue()

    # envoi du rapport en PDF au navigateur
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=magasins.pdf'
    return response

@app.route('/export_magasin_excel')
def export_magasin_excel():
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM magasin")
    data = cur.fetchall()
    cur.close()

    # création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # ajout des titres des colonnes
    headers = [header[0] for header in cur.description]
    for col in range(len(headers)):
        ws.cell(row=1, column=col+1, value=headers[col])

    # ajout des données de la base de données
    for row in range(len(data)):
        for col in range(len(headers)):
            ws.cell(row=row+2, column=col+1, value=data[row][col])

    # conversion du classeur en octets
    output = BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()

    # envoi du classeur Excel au navigateur
    response = make_response(excel_bytes)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=magasins.xlsx'
    return response

#### stock ####

@app.route('/stock', methods=['GET', 'POST'])
def list_stock():
    # Récupération de la page actuelle depuis l'URL
    page = request.args.get(get_page_parameter(), type=int, default=1)

    # Nombre d'éléments à afficher par page
    per_page = 2

    # Définition des filtres
    id_produit = request.form.get('id_produit')
    id_magasin = request.form.get('id_magasin')
    produit = get_nom_produit(id_produit) if id_produit else "Choose"
    magasin = get_nom_magasin(id_magasin) if id_magasin else "Choose"

    # Construction de la requête SQL en fonction des filtres
    query = "SELECT COUNT(*) FROM produit JOIN stock ON produit.id = stock.id_produit JOIN magasin ON magasin.id = stock.id_magasin"
    condition = ""
    if id_produit and id_magasin:
        condition = " WHERE produit.id = %s AND magasin.id = %s"
        query += condition
        params = [id_produit, id_magasin]
        produit = get_nom_produit(id_produit)
        magasin = get_nom_magasin(id_magasin)
    elif id_produit:
        condition = " WHERE produit.id = %s"
        query += condition
        params = [id_produit]
        produit = get_nom_produit(id_produit)
        magasin = "Choose"
    elif id_magasin:
        condition = " WHERE magasin.id = %s"
        query += condition
        params = [id_magasin]
        produit = "Choose"
        magasin = get_nom_magasin(id_magasin)
    else:
        params = []

    # Récupération du nombre total d'éléments dans la table 'stock' en fonction des filtres
    cur = mysql.connection.cursor()
    cur.execute(query, params)
    total_count = cur.fetchone()[0]

    # Calcul de l'offset à partir de la page et du nombre d'éléments par page
    offset = (page - 1) * per_page

    # Construction de la requête SQL pour récupérer les éléments à afficher en fonction des filtres
    query = "SELECT stock.id, produit.nom, produit.prix, magasin.nom, stock.quantite FROM produit JOIN stock ON produit.id = stock.id_produit JOIN magasin ON magasin.id = stock.id_magasin"
    if condition:
        query += condition
    query += " LIMIT %s OFFSET %s"
    params = params + [per_page, offset]

    # Récupération des éléments à afficher pour la page actuelle en fonction des filtres
    cur.execute(query, params)
    stock_info = cur.fetchall()

    # Configuration de la pagination avec le nombre total d'éléments et le nombre d'éléments par page
    pagination = Pagination(page=page, total=total_count, per_page=per_page, css_framework='bootstrap5')

    cur.execute("SELECT * FROM produit")
    product_info = cur.fetchall()

    cur.execute("SELECT * FROM magasin")
    store_info = cur.fetchall()

    cur.close()

    # Rendu du template avec les données récupérées et la pagination
    return render_template('list_stock.html', products=product_info, stores=store_info, otto=stock_info, pagination=pagination, produit=produit, magas=magasin)

@app.route('/add_stock', methods=["POST"])
def add_stock():

    # Obtenir les données du formulaire
    id_produit = request.form['id_produit']
    id_magasin = request.form['id_magasin']
    quantite = request.form['quantite']

    # Établir une connexion à la base de données
    cur = mysql.connection.cursor()

    # Insérer les données dans la table 'stock'
    cur.execute("INSERT INTO stock (id_produit, id_magasin, quantite) VALUES (%s, %s, %s)", (id_produit, id_magasin, quantite))

    # Valider la transaction
    mysql.connection.commit()

    # Fermer la connexion à la base de données
    cur.close()

    # Rediriger vers la page 'magasin'
    return redirect(url_for('list_stock'))

@app.route('/upd_stock/<int:id>', methods=["GET", "POST"])
def upd_stock(id):

    cur = mysql.connection.cursor()
    cur.execute("SELECT stock.id, produit.nom, produit.prix, magasin.nom, stock.quantite FROM produit JOIN stock ON produit.id = stock.id_produit JOIN magasin ON magasin.id = stock.id_magasin WHERE stock.id=%s", (id,))
    stock_info = cur.fetchone()
    cur.close()

    if request.method == "POST":

        quantite = int(request.form['quantite'])# Conversion de la quantité en entier
        # Récupération de la quantité actuelle dans la base de données
        cur = mysql.connection.cursor()
        cur.execute("SELECT quantite FROM stock WHERE id=%s", (id,))
        current_quantite = cur.fetchone()[0]
        # Calcul de la nouvelle quantité
        new_quantite = current_quantite + quantite
        # Mise à jour de la base de données avec la nouvelle quantité
        cur.execute("UPDATE stock SET quantite=%s WHERE id=%s", (new_quantite, id))
        mysql.connection.commit()
        cur.close()
        return redirect(url_for('list_stock'))

    return render_template('upd_stock.html', stock_info=stock_info)


@app.route('/export_stock_pdf')
def export_stock_pdf():
    cur = mysql.connection.cursor()
    cur.execute("SELECT stock.id, produit.nom, produit.prix, magasin.nom, stock.quantite FROM produit JOIN stock ON produit.id = stock.id_produit JOIN magasin ON magasin.id = stock.id_magasin")
    data = cur.fetchall()
    cur.close()

    # création du rapport en PDF
    pdf = FPDF()
    pdf.add_page()

    # ajout de l'en-tête
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Liste des stocks', 0, 1, 'C')
    pdf.set_font('Arial', '', 10)
    pdf.cell(0, 10, 'Ce document contient la liste de tous les stocks', 0, 1, 'C')

    # ajout des titres des colonnes
    pdf.set_font('Arial', 'B', 16)
    col_width = pdf.w / 5.5
    row_height = pdf.font_size * 2
    for header in cur.description:
        pdf.cell(col_width, row_height, str(header[0]), border=1)

    # ajout des données de la base de données
    pdf.set_font('Arial', '', 10)
    for row in data:
        pdf.ln()
        for item in row:
            pdf.cell(col_width, row_height, str(item), border=1)

    # conversion du rapport en PDF en octets
    output = BytesIO()
    pdf.output(output)
    pdf_bytes = output.getvalue()

    # envoi du rapport en PDF au navigateur
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=stocks.pdf'
    return response

@app.route('/export_stock_excel')
def export_stock_excel():
    cur = mysql.connection.cursor()
    cur.execute("SELECT stock.id, produit.nom, produit.prix, magasin.nom, stock.quantite FROM produit JOIN stock ON produit.id = stock.id_produit JOIN magasin ON magasin.id = stock.id_magasin")
    data = cur.fetchall()
    cur.close()

    # création du classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # ajout des titres des colonnes
    headers = [header[0] for header in cur.description]
    for col in range(len(headers)):
        ws.cell(row=1, column=col+1, value=headers[col])

    # ajout des données de la base de données
    for row in range(len(data)):
        for col in range(len(headers)):
            ws.cell(row=row+2, column=col+1, value=data[row][col])

    # conversion du classeur en octets
    output = BytesIO()
    wb.save(output)
    excel_bytes = output.getvalue()

    # envoi du classeur Excel au navigateur
    response = make_response(excel_bytes)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=stocks.xlsx'
    return response

#########################################################################

@app.route('/import_magasin_excel', methods=['POST'])
def import_magasin_excel():
    if 'fichier_excel' not in request.files:
        flash("Aucun fichier sélectionné")
        return redirect(url_for('listemagasin'))
    
    fichier = request.files['fichier_excel']
    
    # Lecture du fichier Excel
    wb = openpyxl.load_workbook(fichier)
    ws = wb.active

    # Extraction des données
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    # Connexion à la base de données
    cur = mysql.connection.cursor()

    # Insertion des données dans la table SQL
    for row in data:
        query = "INSERT INTO magasin (nom, adresse) VALUES (%s, %s)"
        cur.execute(query, (row[1], row[2]))

    # Validation et enregistrement des modifications dans la base de données
    mysql.connection.commit()
    cur.close()
    flash("Importation réussie")
    return redirect(url_for('listemagasin'))



@app.route('/import_produit_excel', methods=['POST'])
def import_produit_excel():
    if 'fichier_excel' not in request.files:
        flash("Aucun fichier sélectionné")
        return redirect(url_for('list_produit'))

    fichier = request.files['fichier_excel']

    # Lecture du fichier Excel
    wb = openpyxl.load_workbook(fichier)
    ws = wb.active

    # Extraction des données
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    # Connexion à la base de données
    cur = mysql.connection.cursor()

    # Insertion des données dans la table SQL
    for row in data:
        query = "INSERT INTO produit (nom, prix, description) VALUES (%s, %s, %s)"
        cur.execute(query, (row[1], row[2], row[3]))

    # Validation et enregistrement des modifications dans la base de données
    mysql.connection.commit()
    cur.close()

    flash("Importation réussie")
    return redirect(url_for('list_produit'))


# Correction de la fonction de récupération du nom de produit et de magasin
def get_nom_produit(id_produit):
    if id_produit:
        cur = mysql.connection.cursor()
        cur.execute("SELECT nom FROM produit WHERE id = %s", [id_produit])
        nom_produit = cur.fetchone()[0]
        cur.close()
        return nom_produit
    else:
        return ""

def get_nom_magasin(id_magasin):
    if id_magasin:
        cur = mysql.connection.cursor()
        cur.execute("SELECT nom FROM magasin WHERE id = %s", [id_magasin])
        nom_magasin = cur.fetchone()[0]
        cur.close()
        return nom_magasin
    else:
        return ""

if __name__ == "__main__":
    app.run(debug=True)